from openpyxl import Workbook, load_workbook 
wb=load_workbook('tests/test1.xlsx')  
ws=wb.active  // uztaisa kjipa failu kur dara darbus(phantom darbs)
Above_3000 = 0  
max_row=ws.max_row  // ieliek row sakitu phantom darba

for i in range (2,max_row+1): // sakot no otras kolonas un otro row
    hours=ws['C' + str(i)].value
    rate=ws['B' + str(i)].value
    if (type(hours)!=str and type(rate)!=str): // pieliek vertibu pie dotajiem
        salary=rate*hours
        ws['D'+str(i)].value=salary // nakosaja rinda ieliek resultatu
        
for i in range (2,max_row+1):
    salary=ws['D'+str(i)].value
    if isinstance(salary,(int, float)) and salary >3000:
        ws['E' +str(i)].value='good'
        Above_3000 +=1
      
print(Above_3000)
wb.save('result.xlsx')
wb.close()


import pandas
get_info=input()  #input information from terminal

fails = pandas.read_excel("description.xlsx", sheet_name="LookupAREA") # if no pages are specified, then the last one saved is open
info_list = fails.values.tolist()

Regions = None 
for row in info_list:  
    if row[1] ==get_info: 
        Regions = row[0]
reg = 0
with open("data.csv") as a:
    next(a)
    for line in a:
        row=line.rstrip().rsplit(",") // atdala rindu no komatiem un ( ) 
        vieta=str(row[1])  // sadala rindu pa dalam 
        summa=int(row[3])
        if vieta==Regions:
            reg+=summa
print(reg)
